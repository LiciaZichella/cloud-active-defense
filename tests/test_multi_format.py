import os
import sys
import types
import itertools
import tempfile
from unittest.mock import patch, MagicMock

_mock_config = types.ModuleType('config')
_mock_config.CONFIG = {
    'localstack': {
        'endpoint': 'http://localhost:4566',
        'access_key': 'test',
        'secret_key': 'test',
        'region': 'us-east-1',
    },
    'buckets': {'documents': 'test-docs'},
    'radar': {'url': 'http://localhost:8080'},
}
sys.modules['config'] = _mock_config
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

with patch('boto3.client') as _mock_boto:
    _mock_boto.return_value.create_bucket.return_value = {}
    import multi_format
    import generator
    import camouflage

from docx import Document
import openpyxl


def test_docx_honey_senza_includepicture():
    with tempfile.TemporaryDirectory() as tmpdir:
        with patch('os.path.dirname', return_value=tmpdir), \
             patch.object(multi_format.s3, 'upload_file'):
            multi_format.crea_documento_docx(
                'honey.docx', 'Titolo', 'Riga1\nRiga2',
                'http://radar/test?file_id=HONEY_abc', 'HONEY', 'Test Autore'
            )
        doc = Document(os.path.join(tmpdir, 'honey.docx'))
        assert 'INCLUDEPICTURE' not in doc.element.xml


def test_docx_real_con_includepicture():
    beacon = 'http://radar/test?file_id=REAL_xyz'
    with tempfile.TemporaryDirectory() as tmpdir:
        with patch('os.path.dirname', return_value=tmpdir), \
             patch.object(multi_format.s3, 'upload_file'):
            multi_format.crea_documento_docx(
                'real.docx', 'Titolo', 'Riga1\nRiga2',
                beacon, 'REAL', 'Test Autore'
            )
        doc = Document(os.path.join(tmpdir, 'real.docx'))
        xml = doc.element.xml
        assert 'INCLUDEPICTURE' in xml
        assert beacon in xml


def test_xlsx_honey_senza_hyperlink():
    with tempfile.TemporaryDirectory() as tmpdir:
        with patch('os.path.dirname', return_value=tmpdir), \
             patch.object(multi_format.s3, 'upload_file'):
            multi_format.crea_documento_xlsx(
                'honey.xlsx', 'Titolo Report', 'Riga1\nRiga2',
                'http://radar/test?file_id=HONEY_abc', 'HONEY', 'Test Autore'
            )
        wb = openpyxl.load_workbook(os.path.join(tmpdir, 'honey.xlsx'))
        ws = wb.active
        assert ws['A1'].value == 'Titolo Report'


def test_xlsx_real_con_hyperlink():
    beacon = 'http://radar/test?file_id=REAL_xyz'
    with tempfile.TemporaryDirectory() as tmpdir:
        with patch('os.path.dirname', return_value=tmpdir), \
             patch.object(multi_format.s3, 'upload_file'):
            multi_format.crea_documento_xlsx(
                'real.xlsx', 'Titolo Report', 'Riga1\nRiga2',
                beacon, 'REAL', 'Test Autore'
            )
        wb = openpyxl.load_workbook(os.path.join(tmpdir, 'real.xlsx'))
        ws = wb.active
        assert 'HYPERLINK' in str(ws['A1'].value or '')
        assert beacon in str(ws['A1'].value or '')


def test_lotto_misto_formati():
    nomi = []

    # Cicliamo le categorie in modo deterministico per coprire tutti e 3 i formati
    _cat_cycle = itertools.cycle(camouflage.CATEGORIE)
    _real_choice = __import__('random').choice

    def _forced_choice(x):
        if x == camouflage.CATEGORIE:
            return next(_cat_cycle)
        return _real_choice(x)

    def cattura(nome_file, *args, **kwargs):
        nomi.append(nome_file)

    with patch('generator.random.choice', side_effect=_forced_choice), \
         patch.object(generator, 'crea_e_carica_documento', side_effect=cattura), \
         patch.object(multi_format, 'crea_documento_docx', side_effect=cattura), \
         patch.object(multi_format, 'crea_documento_xlsx', side_effect=cattura):
        generator.genera_lotto(3, 3)

    assert len(nomi) == 6
    estensioni = {os.path.splitext(n)[1] for n in nomi}
    assert len(estensioni) >= 2
